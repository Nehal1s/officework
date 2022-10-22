# Python program to read an excel file

# import openpyxl module
import openpyxl
import datetime




# Veeva sfs 
_data_rows = ['VEEVA_PARAM',]

x = datetime.datetime.now()

x = x.strftime("%I:%M %p")

# Give the location of the file
path = "VEEVA_IF_RUN_STATUS  15-10-2022 .xlsx"

# To open the workbook
# workbook object is created
_wb = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
_sheet = _wb.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
# VEEVA_FSA_DATA_TRNSMSN_USR_RL_IU	49:43.9	50:54.2

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
for i in range(1, _sheet.max_row + 1):
    _val = _sheet.cell(i, 1).value
    if _val is None:
        break
    _sheet.cell(i, 2).value = x
    _sheet.cell(i, 3).value = x
    print(_val)
    # if _val == 'VEEVA_FSA_DATA_TRNSMSN_USR_RL_IU':
    #     print('found')
    #     print(_sheet.cell(i, 2).value, _sheet.cell(i, 3).value)
    #     break 

_wb.save(path)
# Print value of cell object
# using the value attribute
