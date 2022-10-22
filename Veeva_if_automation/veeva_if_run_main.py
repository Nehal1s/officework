from datetime import datetime
from http import client
import boto3 as boto
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime


# low level client representation
client = boto.client('stepfunctions')


def getDateTime(arn):

    response = client.list_executions(
        stateMachineArn=arn,
        # statusFilter='RUNNING'|'SUCCEEDED'|'FAILED'|'TIMED_OUT'|'ABORTED',
        # only want succeeded and failed onces
        statusFilter='SUCCEEDED' | 'FAILED',
        # Just want the first recent execution
        maxResults=1,
        # nextToken='string'
    )

    _executionArn = response['executions'][0]['executionArn']

    # Getting the reference to recent first execution of ith statemachine
    _sfLast = client.describe_execution(
        executionArn=_executionArn
    )

    _status = _sfLast['status']
    _startDate = _sfLast['startDate']
    _endDate = _sfLast['stopDate']
    return {
        'status': _status,
        'startDate': _startDate,
        'endDate': _endDate
    }


# Arn prefix
_devArnPrefix = 'arn:aws:states:ap-northeast-1:267385292102:stateMachine:'
_ppdArnPrefix = 'arn:aws:states:ap-northeast-1:408643729837:stateMachine:'


# veeva_jobs
_sfs = ['VEEVA_PARAM',
        'VEEVA_CMN_OBJCTS',
        'VEEVA_TSA_CALL',
        'VEEVA_FOR_FSN',
        'VEEVA_IF_LC-KAM',
        'VEEVA_MR_ACTVTY',
        'VEEVA_USR_INFO_FCLTY_DCTR',
        'VEEVA_PRSCRPTN_PATIENTS',
        'VEEVA_ACCNT_PLN_TAGBRL',
        'VEEVA_RFRSH_BIOGRPHY',
        'VEEVA_IF_MSLVEEVA',
        'VEEVA_SURVEY_MR_MDCL',
        'VEEVA_IMF_HCA_STTS',
        'VEEVA_PTNT_RFRL_DSTNTN_RCPNT',
        'VEEVA_TRRTRY_FLD',
        'VEEVA_PRDCT_RLTD',
        'VEEVA_SMPL_PHRMA_PRDCT',
        'VEEVA_AE_SNT_EML',
        'VEEVA_KEY_MSG_CLM',
        'VEEVA_TAG',
        'VEEVA_CPTR_WTHOUT_SBSQNT_AGRGTN',
        'VEEVA_FSA_DATA_MVOUT_DLTN',
        'VEEVA_ZEN',
        'VEEVA_FSA_DATA_TRNSMSN_AZM_BRNCH_OFC',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_BR',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_SECT',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_MR',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_LEVEL_04',
        'VEEVA_FSA_DATA_TRNSMSN_USR_RL_IU',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_LEVEL_04_D',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_MR_D',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_SECT_D',
        'VEEVA_FSA_DATA_TRNSMSN_USR_ROLE_BR_D',
        'VEEVA_FSA_DATA_TRNSMSN_PRDCT',
        'VEEVA_FSA_DATA_TRNSMSN_PRDCT_GRP',
        'VEEVA_TRTRY2_BR_IM',
        'VEEVA_TRTRY2_SECT_IM',
        'VEEVA_TRTRY2_MR_IM',
        'VEEVA_TRTRY2_LEVEL04_IM',
        'VEEVA_TRTRY2_LEVEL04_D',
        'VEEVA_TRTRY2_MR_D',
        'VEEVA_TRTRY2_SECT_D',
        'VEEVA_TRTRY2_BR_D',
        'VEEVA_UT2A',
        'VEEVA_FSA_DATA_TRNSMSN_ACNT',
        'VEEVA_FSA_DATA_TRNSMSN_AE_ACNT',
        'VEEVA_FSA_DATA_ADR_HCP_PRSCPTN',
        'VEEVA_FSA_DATA_TRNSMSN_MSTR_PI',
        'VEEVA_FSA_DATA_TRNSMSN_PRD_PLN_PI',
        'VEEVA_FSA_DATA_TRNSMSN_PRD_STRTGY_PI',
        'VEEVA_FSA_DATA_PRDCT_MTRC_IUD',
        'VEEVA_FSA_DATA_ADPTN_IU',
        'VEEVA_FSA_DATA_ADPTN_D',
        'VEEVA_FSA_DATA_TSF_U',
        'VEEVA_FSA_DATA_TSF_MR_DL_U',
        'VEEVA_FSA_DATA_TSF_NM_U',
        'VEEVA_FSA_DATA_TSF_MY_CNTCT_U',
        'VEEVA_FSA_ACTVTY_DEADLN_JDGMNT_DLY',
        'VEEVA_FSA_ACTVTY_DEADLN_JDGMNT_MNTHLY',
        'VEEVA_FSA_DATA_PRDCT_MTRC_TSFDSW',
        'VEEVA_FSA_JDGMNT_CLSNG_DT_PTNT',
        'VEEVA_PI_TRGT_TRNSMSN',
        'VEEVA_TRTMNT_PLCY_DLY_TRNSMSSN',
        'VEEVA_TRTMNT_PLCY_CRNT_UPDT',
        'VEEVA_TRTMNT_PLCY_PTNTL_RCPTN',
        'VEEVA_TRTMNT_PLCY_NXT_TRM_PTNTL_UPDT',
        'VEEVA_BSNS_DLY_ALWNC',
        'VEEVA_SND_EVNT_ATTNDEE',
        'VEEVA_CONF_HCP_STTS',
        'VEEVA_FCLTY_MR_INFRMTN_TRNSMSN',
        'VEEVA_PRMSN_SET_ASGN',
        'SF_Job_Generic_Outbound_MediChannel_Saccess'
        ]


# workbook creation

x = datetime.datetime.now()
name = "VEEVA_IF_RUN_STATUS " + x.strftime('%d-%b-%Y') + '.xlsx'

wb = Workbook()

ws = wb.active
ws.title = 'VEEVA_IF_RUN_STATUS'

redFill = PatternFill(start_color='BDD7EE',
                      end_color='BDD7EE',
                      fill_type='solid')

fontgreenFill = Font(name='Calibri',
                     size=14,
                     color='1E8900')
fontredFill = Font(name='Calibri',
                   size=14,
                   color='ff0000')


_widths = [63, 22, 22, 22, 22, 43, 53, 56]
_header = ['Batch Name', 'Start time', 'End Time', 'Duration',
           'Status', 'Failure reason', 'Error table count',	'Comments']
alpha = 'A'
for col, val in enumerate(_header):
    _cell = ws.cell(1, col+1)
    _cell.value = val
    _cell.fill = redFill
    ws.column_dimensions[alpha].width = _widths[col]
    alpha = chr(ord(alpha) + 1)


# sf naming
for row, _sf in enumerate(_sfs):
    _arn = _ppdArnPrefix + _sf
    _stts_dict = getDateTime(_arn)
    ws.cell(row + 2, 2).value = _stts_dict['startDate']
    ws.cell(row + 2, 3).value = _stts_dict['endDate']
    ws.cell(row + 2, 4).value = ""
    ws.cell(row + 2, 5).font = fontgreenFill
    if _stts_dict['status']:
        ws.cell(row + 2, 5).font = fontredFill
    ws.cell(row + 2, 5).value = _stts_dict['status']

    # Getting all execution of ith state machine

    # {
    #     'executionArn': 'string',
    #     'stateMachineArn': 'string',
    #     'name': 'string',
    #     'status': 'RUNNING'|'SUCCEEDED'|'FAILED'|'TIMED_OUT'|'ABORTED',
    #     'startDate': datetime(2015, 1, 1),
    #     'stopDate': datetime(2015, 1, 1),
    #     'input': 'string',
    #     'inputDetails': {
    #         'included': True|False
    #     },
    #     'output': 'string',
    #     'outputDetails': {
    #         'included': True|False
    #     },
    #     'traceHeader': 'string'
    # }


wb.save(filename=name)
