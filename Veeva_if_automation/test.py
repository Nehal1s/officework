from tkinter import font
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime

x = datetime.datetime.now()
name = "VEEVA_IF_RUN_STATUS " + x.strftime('%d-%b-%Y') + '.xlsx'

wb = Workbook()

ws = wb.active
ws.title = 'VEEVA_IF_RUN_STATUS'

redFill = PatternFill(start_color='BDD7EE',
                      end_color='BDD7EE',
                      fill_type='solid')
fontFill = Font(name='Calibri',
                size=14,
                color='1E8900')


_widths = [63, 22, 22, 22, 22, 43, 53, 56]
_header = ['Batch Name', 'Start time', 'End Time', 'Duration',
           'Status', 'Failure reason', 'Error table count',	'Comments']
alpha = 'A'
for col, val in enumerate(_header):
    _cell = ws.cell(1, col+1)
    _cell.value = val
    # _cell.fill = redFill
    # _cell.font = fontFill
    ws.column_dimensions[alpha].width = _widths[col]
    alpha = chr(ord(alpha) + 1)


_file = wb.save(filename=name)
print(_file)